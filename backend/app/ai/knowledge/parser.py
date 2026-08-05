"""文档解析器：PDF/Word/Excel → 文本段落"""
from pathlib import Path

async def parse_pdf(file_path: str) -> str:
    """解析 PDF 文件，返回文本内容"""
    try:
        from PyPDF2 import PdfReader
        reader = PdfReader(file_path)
        text = ""
        for page in reader.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"
        return text.strip()
    except ImportError:
        return "PDF 解析需要安装 PyPDF2 库"
    except Exception as e:
        return f"PDF 解析失败: {str(e)}"

async def parse_word(file_path: str) -> str:
    """解析 Word 文件"""
    try:
        from docx import Document
        doc = Document(file_path)
        text = "\n".join([p.text for p in doc.paragraphs if p.text.strip()])
        return text.strip()
    except ImportError:
        return "Word 解析需要安装 python-docx 库"
    except Exception as e:
        return f"Word 解析失败: {str(e)}"

async def parse_excel(file_path: str) -> str:
    """解析 Excel 文件，每行转为文本段落"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        all_text = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_text = []
            for row in ws.iter_rows(values_only=True):
                row_str = " | ".join([str(cell) for cell in row if cell is not None])
                if row_str.strip():
                    rows_text.append(row_str)
            if rows_text:
                all_text.append(f"[Sheet: {sheet_name}]\n" + "\n".join(rows_text))
        return "\n\n".join(all_text).strip()
    except ImportError:
        return "Excel 解析需要安装 openpyxl 库"
    except Exception as e:
        return f"Excel 解析失败: {str(e)}"

async def parse_document(file_path: str, file_type: str) -> str:
    """统一入口：根据文件类型调用对应解析器"""
    parsers = {
        "pdf": parse_pdf,
        "word": parse_word,
        "excel": parse_excel,
        "txt": parse_txt_sync,
    }
    parser = parsers.get(file_type)
    if not parser:
        raise ValueError(f"不支持的文件类型: {file_type}")
    return await parser(file_path)


async def parse_txt_sync(file_path: str) -> str:
    """TXT 文件读取（同步IO包在 async 中）"""
    with open(file_path, "r", encoding="utf-8") as f:
        return f.read()
